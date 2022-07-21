import openpyxl as xl
import pandas as pd
import xlsxwriter
import glob
#import os
from tkinter import filedialog
from tkinter import *

# GUI
def browse_button1():
    # Allow user to select a directory and store it in global var
    # called folder_path
    global folder_path
    filename = filedialog.askdirectory()
    folder_path.set(filename)
    #print(filename)
def browse_button2():
    # Allow user to select a directory and store it in global var
    # called folder_path
    global folder_path2
    filename = filedialog.askdirectory()
    folder_path2.set(filename)
    #print(filename)
root = Tk()
root.title('Project Finance Hour Plan Generator')
# Browse Input Folder
folder_path = StringVar()
lbl1 = Label(master=root,textvariable=folder_path)
lbl1.grid(row=0, column=3)
button2 = Button(text="Browse Input Folder", command=browse_button1)
button2.grid(row=0, column=1)
# Browse Output Folder
folder_path2 = StringVar()
lbl1 = Label(master=root,textvariable=folder_path2)
lbl1.grid(row=1, column=3)
button2 = Button(text="Browse Output Folder", command=browse_button2)
button2.grid(row=1, column=1)
# Generate button
button2 = Button(text="Generate", command=browse_button2)
button2.grid(row=2, column=1)
mainloop()

src_path = folder_path.get()
dest_path = folder_path2.get()

dest_file_list = ['Hour Summary_Plan&Actual.xlsx','Hour Summary_ByMember.xlsx']
dest_path_list = []
for dest_file in dest_file_list:
    dest_path_list.append(str(dest_path)+'\\'+dest_file)
file_path_list = glob.glob(str(src_path)+'\*.xlsx')
for dest_file in dest_file_list:
    for file_path in file_path_list:
        if dest_file in file_path:
            file_path_list.remove(file_path)

#create the dataframe
df_output = pd.DataFrame(columns=['Project Name','Name','Week','Hour','Cost'])

#create one new excel as destination
for dest_path in dest_path_list:
    workbook  = xlsxwriter.Workbook(dest_path)
    workbook.close()

# opening the destination excel file
wb_planandact = xl.load_workbook(filename=dest_path_list[0])
wb_bymember = xl.load_workbook(filename=dest_path_list[1])

for file_path in file_path_list:
    # ==========Merge Summary View==========
    # opening the source excel file
    wb1 = xl.load_workbook(filename=file_path, data_only=True)
 
    ws1 = wb1["Summary View"]
    # create sheet in dest excel
    file_name = file_path.split("\\")[-1].split(".")[0]
    wb_planandact.create_sheet(file_name)
    ws2 = wb_planandact[file_name]
    # copy to dest excel
    for row in ws1:
        for cell in row:
            ws2[cell.coordinate] = cell.value
    
    #==========Merge Plan Hour Week==========
    #initial the parameters
    project_name = ''
    name_list = []
    week_list = []
    hour = []
    cost = []
    #gather data in original workbooks
    #turn into dataframe
    df = pd.read_excel(file_path, sheet_name='PlanHoursWeek')
    #get column list
    column_list = df.columns.values.tolist()
    #get project_name
    project_name = column_list[0]
    #get name_list
    i = 1
    while i < len(column_list):
        if 'Unnamed' not in column_list[i] and 'Sub Total' not in column_list[i]:
            name_list.append(column_list[i])
        i = i + 2
    #get week_list
    week_list = df[column_list[0]].values.tolist()[2:-2]
    for i in range(len(week_list)):
        week_list[i] = week_list[i].strftime('%Y-%m-%d')
    #get hour
    for name in name_list:
        temp_hour = df[name].values.tolist()[2:-2]
        hour.append(temp_hour)
    #cost
    for name in name_list:
        temp_cost = df[column_list[column_list.index(name)+1]].values.tolist()[2:-2]
        cost.append(temp_cost)
    #put data into dataframe
    for name_idx in range(len(name_list)):
        for week_idx in range(len(week_list)):
            df_output.loc[df_output.shape[0]] = [project_name,name_list[name_idx],week_list[week_idx],hour[name_idx][week_idx],cost[name_idx][week_idx]]
            
#turn dataframe to pivot table
df_output1 = df_output.pivot_table(index=['Name', 'Project Name'], columns='Week', values='Hour').fillna(0)
df_output2 = df_output[['Name','Week','Hour']].groupby(['Name','Week'],as_index = False).sum().pivot('Name','Week').fillna(0)

#save the dataframe to exist excel    
writer = pd.ExcelWriter(dest_path_list[1], engine='openpyxl') 
writer.book = wb_bymember
df_output1.to_excel(writer, "Member Weekly Plan Hr by Proj")
df_output2.to_excel(writer, "Member Weekly Plan TTL Hr")
writer.save()

writer = pd.ExcelWriter(dest_path_list[1], engine = 'xlsxwriter')
df_output1.to_excel(writer, "Member Weekly Plan Hr by Proj")
df_output2.to_excel(writer, "Member Weekly Plan TTL Hr")
writer.save()

#remove default sheet
del wb_planandact['Sheet1']
del wb_bymember['Sheet1']
wb_planandact.save(dest_path_list[0])
wb_planandact.close()
wb_bymember.close()